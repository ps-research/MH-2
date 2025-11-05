"""
Administrative operations for reset, archival, and data management.
"""
import os
import json
import shutil
import tarfile
import hashlib
import logging
from pathlib import Path
from typing import Dict, Optional, List
from datetime import datetime
import redis
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

from ..core.checkpoint import RedisCheckpointManager
from ..storage.excel_manager import ExcelAnnotationManager
from ..storage.malform_logger import MalformLogger
from ..workers.controller import WorkerController


logger = logging.getLogger(__name__)


class AdminOperations:
    """
    Administrative operations for system management.

    Features:
    - Reset operations (domain, annotator, run, factory)
    - State export/import
    - Data archival with compression
    - Excel file consolidation
    - Audit logging
    """

    def __init__(self, redis_client: redis.Redis):
        """
        Initialize admin operations.

        Args:
            redis_client: Redis client instance
        """
        self.redis = redis_client
        self.checkpoint_mgr = RedisCheckpointManager(redis_client)
        self.excel_mgr = ExcelAnnotationManager(
            output_dir='data/annotations',
            redis_client=redis_client
        )
        self.malform_logger = MalformLogger(
            log_dir='data/malform_logs',
            redis_client=redis_client
        )
        self.controller = WorkerController(redis_client)

        # Archive directory
        self.archive_dir = Path('data/archive')
        self.archive_dir.mkdir(parents=True, exist_ok=True)

        # Audit log
        self.audit_log_path = Path('data/admin_audit.log')

        logger.info("AdminOperations initialized")

    # ═══════════════════════════════════════════════════════════
    # AUDIT LOGGING
    # ═══════════════════════════════════════════════════════════

    def _log_audit(self, operation: str, details: Dict) -> None:
        """
        Log administrative operation to audit log.

        Args:
            operation: Operation name
            details: Operation details
        """
        audit_entry = {
            'timestamp': datetime.now().isoformat(),
            'operation': operation,
            'details': details
        }

        try:
            with open(self.audit_log_path, 'a') as f:
                f.write(json.dumps(audit_entry) + '\n')

            logger.info(f"Audit log: {operation}")

        except Exception as e:
            logger.error(f"Error writing audit log: {e}")

    # ═══════════════════════════════════════════════════════════
    # RESET OPERATIONS
    # ═══════════════════════════════════════════════════════════

    def reset_domain(
        self,
        annotator_id: int,
        domain: str,
        keep_excel: bool = False
    ) -> Dict:
        """
        Reset a specific annotator-domain pair.

        Steps:
        1. Stop affected worker
        2. Clear Redis checkpoints
        3. Clear malform logs
        4. Handle Excel file (delete or archive)
        5. Clear task queue

        Args:
            annotator_id: Annotator ID
            domain: Domain name
            keep_excel: If True, archive Excel file instead of deleting

        Returns:
            Dictionary with reset results
        """
        worker_key = f"{annotator_id}_{domain}"

        logger.warning(f"Resetting domain: {worker_key} (keep_excel={keep_excel})")

        result = {
            'operation': 'reset_domain',
            'worker_key': worker_key,
            'keep_excel': keep_excel,
            'timestamp': datetime.now().isoformat(),
            'steps': {}
        }

        try:
            # Step 1: Stop worker
            try:
                self.controller.stop_worker(annotator_id, domain, force=False)
                result['steps']['stop_worker'] = 'SUCCESS'
            except Exception as e:
                result['steps']['stop_worker'] = f'FAILED: {e}'
                logger.error(f"Error stopping worker: {e}")

            # Step 2: Clear Redis checkpoint
            try:
                self.checkpoint_mgr.clear_domain(annotator_id, domain)
                result['steps']['clear_checkpoint'] = 'SUCCESS'
            except Exception as e:
                result['steps']['clear_checkpoint'] = f'FAILED: {e}'
                logger.error(f"Error clearing checkpoint: {e}")

            # Step 3: Clear malform logs
            try:
                # Clear Redis malforms
                cleared = self.malform_logger.clear_malforms(annotator_id, domain)

                # Handle JSON file
                malform_file = self.malform_logger._get_file_path(annotator_id, domain)
                if malform_file.exists():
                    if keep_excel:
                        # Archive malform log
                        archive_path = self._archive_file(malform_file, 'malform_logs')
                        result['malform_log_archived'] = str(archive_path)
                    else:
                        malform_file.unlink()

                result['steps']['clear_malforms'] = f'SUCCESS: {cleared} cleared'

            except Exception as e:
                result['steps']['clear_malforms'] = f'FAILED: {e}'
                logger.error(f"Error clearing malforms: {e}")

            # Step 4: Handle Excel file
            try:
                excel_file = self.excel_mgr._get_file_path(annotator_id, domain)

                if excel_file.exists():
                    if keep_excel:
                        # Archive Excel file
                        archive_path = self._archive_file(excel_file, 'annotations')
                        result['excel_archived'] = str(archive_path)
                        result['steps']['handle_excel'] = f'SUCCESS: Archived to {archive_path}'
                    else:
                        excel_file.unlink()
                        result['steps']['handle_excel'] = 'SUCCESS: Deleted'
                else:
                    result['steps']['handle_excel'] = 'SUCCESS: No file found'

            except Exception as e:
                result['steps']['handle_excel'] = f'FAILED: {e}'
                logger.error(f"Error handling Excel file: {e}")

            # Step 5: Clear task queue
            try:
                from ..core.celery_app import app, get_queue_name

                queue_name = get_queue_name(annotator_id, domain)
                # Note: Purging queue requires Celery control access
                # For now, just log it
                result['steps']['clear_queue'] = 'PENDING: Manual purge required'

            except Exception as e:
                result['steps']['clear_queue'] = f'FAILED: {e}'
                logger.error(f"Error clearing queue: {e}")

            result['success'] = True

            # Audit log
            self._log_audit('reset_domain', {
                'annotator_id': annotator_id,
                'domain': domain,
                'keep_excel': keep_excel,
                'result': result
            })

        except Exception as e:
            result['success'] = False
            result['error'] = str(e)
            logger.error(f"Error resetting domain {worker_key}: {e}")

        return result

    def reset_annotator(self, annotator_id: int, keep_excel: bool = False) -> Dict:
        """
        Reset all domains for a specific annotator.

        Args:
            annotator_id: Annotator ID
            keep_excel: If True, archive Excel files

        Returns:
            Dictionary with reset results
        """
        logger.warning(f"Resetting annotator: {annotator_id} (keep_excel={keep_excel})")

        result = {
            'operation': 'reset_annotator',
            'annotator_id': annotator_id,
            'keep_excel': keep_excel,
            'timestamp': datetime.now().isoformat(),
            'domains': {}
        }

        domains = ['urgency', 'therapeutic', 'intensity', 'adjunct', 'modality', 'redressal']

        for domain in domains:
            domain_result = self.reset_domain(annotator_id, domain, keep_excel=keep_excel)
            result['domains'][domain] = domain_result

        result['success'] = all(r.get('success', False) for r in result['domains'].values())

        # Audit log
        self._log_audit('reset_annotator', {
            'annotator_id': annotator_id,
            'keep_excel': keep_excel,
            'result': result
        })

        return result

    def reset_run(self, run_id: str) -> Dict:
        """
        Reset a specific run (not implemented - placeholder).

        Args:
            run_id: Run identifier

        Returns:
            Dictionary with reset results
        """
        logger.warning(f"Resetting run: {run_id}")

        result = {
            'operation': 'reset_run',
            'run_id': run_id,
            'timestamp': datetime.now().isoformat(),
            'message': 'Run-based reset not yet implemented'
        }

        return result

    def factory_reset(self, confirm: bool = False) -> Dict:
        """
        Factory reset - clear ALL data (use with extreme caution!).

        Steps:
        1. Require confirmation
        2. Stop all workers
        3. Archive all data
        4. Flush all Redis databases
        5. Delete/archive all Excel files
        6. Archive all logs
        7. Reinitialize Redis

        Args:
            confirm: Must be True to proceed (safety check)

        Returns:
            Dictionary with reset results
        """
        if not confirm:
            return {
                'operation': 'factory_reset',
                'error': 'Confirmation required. Set confirm=True to proceed.',
                'warning': 'This will DELETE ALL DATA!'
            }

        logger.critical("FACTORY RESET INITIATED!")

        result = {
            'operation': 'factory_reset',
            'timestamp': datetime.now().isoformat(),
            'steps': {},
            'archive_location': None
        }

        try:
            # Create timestamped archive
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archive_name = f"factory_reset_backup_{timestamp}"

            # Step 1: Stop all workers
            try:
                stop_results = self.controller.stop_all(force=False)
                result['steps']['stop_all_workers'] = f"SUCCESS: {sum(stop_results.values())} workers stopped"
            except Exception as e:
                result['steps']['stop_all_workers'] = f'FAILED: {e}'

            # Step 2: Archive all data
            try:
                archive_path = self.archive_data(archive_name)
                result['archive_location'] = archive_path
                result['steps']['archive_data'] = f'SUCCESS: {archive_path}'
            except Exception as e:
                result['steps']['archive_data'] = f'FAILED: {e}'

            # Step 3: Flush all Redis databases
            try:
                self.checkpoint_mgr.factory_reset()
                result['steps']['flush_redis'] = 'SUCCESS'
            except Exception as e:
                result['steps']['flush_redis'] = f'FAILED: {e}'

            # Step 4: Clear all Excel files
            try:
                annotations_dir = Path('data/annotations')
                if annotations_dir.exists():
                    for excel_file in annotations_dir.glob('*.xlsx'):
                        excel_file.unlink()
                result['steps']['clear_excel'] = 'SUCCESS'
            except Exception as e:
                result['steps']['clear_excel'] = f'FAILED: {e}'

            # Step 5: Clear malform logs
            try:
                malform_dir = Path('data/malform_logs')
                if malform_dir.exists():
                    for json_file in malform_dir.glob('*.json'):
                        json_file.unlink()
                result['steps']['clear_malforms'] = 'SUCCESS'
            except Exception as e:
                result['steps']['clear_malforms'] = f'FAILED: {e}'

            # Step 6: Clear worker logs
            try:
                log_dir = Path('data/logs')
                if log_dir.exists():
                    for log_file in log_dir.glob('*.log'):
                        log_file.unlink()
                result['steps']['clear_logs'] = 'SUCCESS'
            except Exception as e:
                result['steps']['clear_logs'] = f'FAILED: {e}'

            result['success'] = True

            # Audit log
            self._log_audit('factory_reset', {'result': result})

            logger.critical(f"FACTORY RESET COMPLETE! Archive: {result['archive_location']}")

        except Exception as e:
            result['success'] = False
            result['error'] = str(e)
            logger.error(f"Error in factory reset: {e}")

        return result

    # ═══════════════════════════════════════════════════════════
    # ARCHIVAL OPERATIONS
    # ═══════════════════════════════════════════════════════════

    def _archive_file(self, file_path: Path, category: str) -> Path:
        """
        Archive a single file.

        Args:
            file_path: Path to file to archive
            category: Category for organizing archives

        Returns:
            Path to archived file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive_subdir = self.archive_dir / f"{category}_{timestamp}"
        archive_subdir.mkdir(parents=True, exist_ok=True)

        dest_path = archive_subdir / file_path.name
        shutil.copy2(file_path, dest_path)

        logger.debug(f"Archived {file_path.name} to {dest_path}")

        return dest_path

    def archive_data(self, archive_name: str, compress: bool = True) -> str:
        """
        Archive all data (Excel files, logs, malform logs, Redis state).

        Args:
            archive_name: Name for archive
            compress: If True, create .tar.gz archive

        Returns:
            Path to archive directory or tarball
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive_name_full = f"{archive_name}_{timestamp}"

        archive_path = self.archive_dir / archive_name_full
        archive_path.mkdir(parents=True, exist_ok=True)

        logger.info(f"Creating archive: {archive_name_full}")

        try:
            # Archive Excel files
            annotations_dir = Path('data/annotations')
            if annotations_dir.exists():
                excel_dest = archive_path / 'annotations'
                excel_dest.mkdir(exist_ok=True)

                for excel_file in annotations_dir.glob('*.xlsx'):
                    shutil.copy2(excel_file, excel_dest / excel_file.name)

            # Archive malform logs
            malform_dir = Path('data/malform_logs')
            if malform_dir.exists():
                malform_dest = archive_path / 'malform_logs'
                malform_dest.mkdir(exist_ok=True)

                for json_file in malform_dir.glob('*.json'):
                    shutil.copy2(json_file, malform_dest / json_file.name)

            # Archive worker logs
            log_dir = Path('data/logs')
            if log_dir.exists():
                logs_dest = archive_path / 'logs'
                logs_dest.mkdir(exist_ok=True)

                for log_file in log_dir.glob('*.log'):
                    shutil.copy2(log_file, logs_dest / log_file.name)

            # Export Redis state
            redis_export = self.export_state(str(archive_path / 'redis_state.json'))

            # Create metadata file
            metadata = {
                'archive_name': archive_name_full,
                'created_at': datetime.now().isoformat(),
                'archived_components': {
                    'excel_files': len(list((archive_path / 'annotations').glob('*.xlsx'))) if (archive_path / 'annotations').exists() else 0,
                    'malform_logs': len(list((archive_path / 'malform_logs').glob('*.json'))) if (archive_path / 'malform_logs').exists() else 0,
                    'worker_logs': len(list((archive_path / 'logs').glob('*.log'))) if (archive_path / 'logs').exists() else 0,
                    'redis_state': redis_export is not None
                }
            }

            # Calculate checksums
            checksums = {}
            for file_path in archive_path.rglob('*'):
                if file_path.is_file():
                    rel_path = file_path.relative_to(archive_path)
                    checksums[str(rel_path)] = self._calculate_checksum(file_path)

            metadata['checksums'] = checksums

            # Write metadata
            with open(archive_path / 'archive_metadata.json', 'w') as f:
                json.dump(metadata, f, indent=2)

            # Compress if requested
            if compress:
                tarball_path = f"{archive_path}.tar.gz"

                with tarfile.open(tarball_path, "w:gz") as tar:
                    tar.add(archive_path, arcname=archive_name_full)

                # Remove uncompressed directory
                shutil.rmtree(archive_path)

                logger.info(f"Archive created and compressed: {tarball_path}")
                return tarball_path

            else:
                logger.info(f"Archive created: {archive_path}")
                return str(archive_path)

        except Exception as e:
            logger.error(f"Error creating archive: {e}")
            raise

    def _calculate_checksum(self, file_path: Path) -> str:
        """Calculate SHA256 checksum for a file."""
        sha256 = hashlib.sha256()

        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b''):
                sha256.update(chunk)

        return sha256.hexdigest()

    # ═══════════════════════════════════════════════════════════
    # STATE EXPORT/IMPORT
    # ═══════════════════════════════════════════════════════════

    def export_state(self, output_path: str) -> Optional[str]:
        """
        Export all Redis state to JSON file.

        Args:
            output_path: Output JSON file path

        Returns:
            Path to exported file or None if failed
        """
        logger.info(f"Exporting Redis state to {output_path}")

        try:
            state = {
                'timestamp': datetime.now().isoformat(),
                'checkpoint_summary': self.checkpoint_mgr.get_summary(),
                'checkpoints': {},
                'progress': {},
                'workers': {},
                'metrics': {}
            }

            # Export checkpoints
            checkpoint_keys = self.redis.keys("checkpoint:*")
            for key in checkpoint_keys:
                completed = list(self.redis.smembers(key))
                state['checkpoints'][key] = completed

            # Export progress
            progress_keys = self.redis.keys("progress:*")
            for key in progress_keys:
                progress = self.redis.hgetall(key)
                state['progress'][key] = dict(progress)

            # Export workers
            worker_keys = self.redis.keys("worker:*")
            for key in worker_keys:
                worker_data = self.redis.hgetall(key)
                state['workers'][key] = dict(worker_data)

            # Export metrics
            metrics_keys = self.redis.keys("metrics:*")
            for key in metrics_keys:
                metrics = self.redis.hgetall(key)
                state['metrics'][key] = dict(metrics)

            # Write to file
            with open(output_path, 'w') as f:
                json.dump(state, f, indent=2)

            logger.info(f"Redis state exported to {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error exporting state: {e}")
            return None

    def import_state(self, snapshot_path: str, merge: bool = False) -> Dict:
        """
        Import Redis state from JSON file.

        Args:
            snapshot_path: Path to snapshot JSON file
            merge: If True, merge with existing state. If False, replace.

        Returns:
            Dictionary with import results
        """
        logger.info(f"Importing Redis state from {snapshot_path} (merge={merge})")

        result = {
            'operation': 'import_state',
            'snapshot_path': snapshot_path,
            'merge': merge,
            'timestamp': datetime.now().isoformat(),
            'imported': {}
        }

        try:
            # Stop all workers first
            self.controller.stop_all(force=False)

            # Load snapshot
            with open(snapshot_path, 'r') as f:
                state = json.load(f)

            # Validate schema
            required_keys = ['checkpoints', 'progress', 'workers']
            if not all(k in state for k in required_keys):
                result['error'] = 'Invalid snapshot: missing required keys'
                return result

            # Clear existing state if not merging
            if not merge:
                self.checkpoint_mgr.factory_reset()

            # Use Redis transaction for atomic restore
            pipe = self.redis.pipeline()

            # Restore checkpoints
            for key, completed_samples in state.get('checkpoints', {}).items():
                if completed_samples:
                    pipe.sadd(key, *completed_samples)
            result['imported']['checkpoints'] = len(state.get('checkpoints', {}))

            # Restore progress
            for key, progress in state.get('progress', {}).items():
                pipe.hset(key, mapping=progress)
            result['imported']['progress'] = len(state.get('progress', {}))

            # Restore workers
            for key, worker_data in state.get('workers', {}).items():
                pipe.hset(key, mapping=worker_data)
            result['imported']['workers'] = len(state.get('workers', {}))

            # Restore metrics
            for key, metrics in state.get('metrics', {}).items():
                pipe.hset(key, mapping=metrics)
            result['imported']['metrics'] = len(state.get('metrics', {}))

            # Execute transaction
            pipe.execute()

            # Verify Excel file integrity
            excel_integrity = self.excel_mgr
            # (Verification would happen here)

            result['success'] = True

            # Audit log
            self._log_audit('import_state', result)

            logger.info("Redis state import complete")

        except Exception as e:
            result['success'] = False
            result['error'] = str(e)
            logger.error(f"Error importing state: {e}")

        return result

    # ═══════════════════════════════════════════════════════════
    # EXCEL CONSOLIDATION
    # ═══════════════════════════════════════════════════════════

    def consolidate_excel_files(self) -> Dict:
        """
        Consolidate all Excel files into a single workbook.

        Each annotator gets one worksheet with all domains combined.
        Adds a summary worksheet with progress statistics.

        Returns:
            Dictionary with consolidation results
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = Path(f'data/consolidated_annotations_{timestamp}.xlsx')

        logger.info(f"Consolidating Excel files to {output_path}")

        result = {
            'operation': 'consolidate_excel',
            'output_path': str(output_path),
            'timestamp': datetime.now().isoformat(),
            'worksheets': {},
            'total_rows': 0
        }

        try:
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # Consolidate by annotator
            for annotator_id in range(1, 6):
                ws = wb.create_sheet(f"Annotator_{annotator_id}")

                # Add headers
                headers = ['Domain', 'Sample_ID', 'Text', 'Raw_Response', 'Label',
                          'Malformed_Flag', 'Parsing_Error', 'Validity_Error', 'Timestamp']
                ws.append(headers)

                # Format headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(1, col_num)
                    cell.fill = header_fill
                    cell.font = header_font

                row_count = 0

                # Add data from all domains
                domains = ['urgency', 'therapeutic', 'intensity', 'adjunct', 'modality', 'redressal']

                for domain in domains:
                    excel_file = self.excel_mgr._get_file_path(annotator_id, domain)

                    if not excel_file.exists():
                        continue

                    try:
                        src_wb = load_workbook(excel_file, read_only=True)
                        src_ws = src_wb.active

                        # Copy data rows (skip header)
                        for row in src_ws.iter_rows(min_row=2, values_only=True):
                            if row[0]:  # Has sample_id
                                # Add domain as first column
                                ws.append([domain] + list(row))
                                row_count += 1

                        src_wb.close()

                    except Exception as e:
                        logger.error(f"Error reading {excel_file.name}: {e}")

                result['worksheets'][f'Annotator_{annotator_id}'] = row_count
                result['total_rows'] += row_count

            # Add summary worksheet
            summary_ws = wb.create_sheet('Summary', 0)  # Insert at beginning

            summary_data = [
                ['Consolidated Annotation Summary'],
                ['Generated:', datetime.now().isoformat()],
                [''],
                ['Annotator', 'Total Annotations']
            ]

            for annotator_id in range(1, 6):
                count = result['worksheets'].get(f'Annotator_{annotator_id}', 0)
                summary_data.append([f'Annotator {annotator_id}', count])

            summary_data.extend([
                [''],
                ['Grand Total', result['total_rows']]
            ])

            for row in summary_data:
                summary_ws.append(row)

            # Format summary
            summary_ws['A1'].font = Font(bold=True, size=14)

            # Save workbook
            wb.save(output_path)

            result['success'] = True

            logger.info(f"Consolidation complete: {result['total_rows']} total rows")

        except Exception as e:
            result['success'] = False
            result['error'] = str(e)
            logger.error(f"Error consolidating Excel files: {e}")

        return result
