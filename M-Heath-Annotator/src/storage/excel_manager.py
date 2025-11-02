"""
Excel annotation manager with file locking and checkpoint synchronization.
"""
import os
import sys
import time
import atexit
import logging
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from datetime import datetime
from contextlib import contextmanager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import redis

from ..core.checkpoint import RedisCheckpointManager


logger = logging.getLogger(__name__)


# Platform-specific file locking
if sys.platform == 'win32':
    import msvcrt

    def lock_file(file_handle):
        """Lock file on Windows."""
        msvcrt.locking(file_handle.fileno(), msvcrt.LK_NBLCK, 1)

    def unlock_file(file_handle):
        """Unlock file on Windows."""
        msvcrt.locking(file_handle.fileno(), msvcrt.LK_UNLCK, 1)
else:
    import fcntl

    def lock_file(file_handle):
        """Lock file on Unix."""
        fcntl.flock(file_handle.fileno(), fcntl.LOCK_EX)

    def unlock_file(file_handle):
        """Unlock file on Unix."""
        fcntl.flock(file_handle.fileno(), fcntl.LOCK_UN)


class ExcelAnnotationManager:
    """
    Manages annotation storage in Excel files with file locking.

    Features:
    - One Excel file per worker (annotator-domain pair)
    - File locking for concurrent access
    - Write buffering and batch commits
    - Checkpoint synchronization
    - Resume capability
    """

    # Excel column headers
    HEADERS = [
        'Sample_ID',
        'Text',
        'Raw_Response',
        'Label',
        'Malformed_Flag',
        'Parsing_Error',
        'Validity_Error',
        'Timestamp'
    ]

    def __init__(
        self,
        output_dir: str,
        redis_client: redis.Redis,
        buffer_size: int = 10
    ):
        """
        Initialize Excel annotation manager.

        Args:
            output_dir: Directory for annotation Excel files
            redis_client: Redis client for checkpoint synchronization
            buffer_size: Number of rows to buffer before flush
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.redis = redis_client
        self.checkpoint_mgr = RedisCheckpointManager(redis_client)
        self.buffer_size = buffer_size

        # Write buffers: {worker_key: [row_data, ...]}
        self._buffers: Dict[str, List[Dict]] = {}

        # Register cleanup on exit
        atexit.register(self.flush_all_buffers)

        logger.info(f"ExcelAnnotationManager initialized: {self.output_dir}")

    def _get_worker_key(self, annotator_id: int, domain: str) -> str:
        """Get unique key for worker."""
        return f"{annotator_id}_{domain}"

    def _get_file_path(self, annotator_id: int, domain: str) -> Path:
        """Get Excel file path for worker."""
        filename = f"annotator_{annotator_id}_{domain}.xlsx"
        return self.output_dir / filename

    @contextmanager
    def lock_file_context(self, file_path: Path, max_retries: int = 5, base_delay: float = 0.5):
        """
        Context manager for file locking with retry.

        Args:
            file_path: Path to file to lock
            max_retries: Maximum lock acquisition retries
            base_delay: Base delay for exponential backoff

        Yields:
            File handle

        Raises:
            IOError: If unable to acquire lock
        """
        for attempt in range(max_retries):
            try:
                file_handle = open(file_path, 'r+b')
                lock_file(file_handle)
                logger.debug(f"Acquired lock on {file_path.name}")

                try:
                    yield file_handle
                finally:
                    unlock_file(file_handle)
                    file_handle.close()
                    logger.debug(f"Released lock on {file_path.name}")

                return

            except (IOError, OSError) as e:
                if attempt < max_retries - 1:
                    delay = base_delay * (2 ** attempt)
                    logger.warning(f"Failed to acquire lock on {file_path.name}, retry in {delay}s: {e}")
                    time.sleep(delay)
                else:
                    logger.error(f"Failed to acquire lock after {max_retries} attempts")
                    raise IOError(f"Unable to lock file: {file_path}")

    def initialize_file(self, annotator_id: int, domain: str) -> str:
        """
        Initialize Excel file for worker if it doesn't exist.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Path to initialized file
        """
        file_path = self._get_file_path(annotator_id, domain)

        if file_path.exists():
            logger.debug(f"File already exists: {file_path.name}")
            return str(file_path)

        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = domain

        # Write headers
        ws.append(self.HEADERS)

        # Format headers
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")

        for col_num, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(1, col_num)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        column_widths = {
            'A': 15,  # Sample_ID
            'B': 50,  # Text
            'C': 50,  # Raw_Response
            'D': 20,  # Label
            'E': 12,  # Malformed_Flag
            'F': 30,  # Parsing_Error
            'G': 30,  # Validity_Error
            'H': 20   # Timestamp
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Save workbook
        wb.save(file_path)
        logger.info(f"Created Excel file: {file_path.name}")

        return str(file_path)

    def write_annotation(self, annotator_id: int, domain: str, row_data: Dict) -> None:
        """
        Write single annotation to Excel file (buffered).

        Args:
            annotator_id: Annotator ID
            domain: Domain name
            row_data: Annotation data dictionary
        """
        worker_key = self._get_worker_key(annotator_id, domain)

        # Initialize buffer if needed
        if worker_key not in self._buffers:
            self._buffers[worker_key] = []

        # Add to buffer
        self._buffers[worker_key].append(row_data)

        # Flush if buffer full
        if len(self._buffers[worker_key]) >= self.buffer_size:
            self.flush_buffer(annotator_id, domain)

    def batch_write(self, annotator_id: int, domain: str, rows: List[Dict]) -> None:
        """
        Write multiple annotations at once.

        Args:
            annotator_id: Annotator ID
            domain: Domain name
            rows: List of row data dictionaries
        """
        if not rows:
            return

        # Initialize file if needed
        file_path = self._get_file_path(annotator_id, domain)
        if not file_path.exists():
            self.initialize_file(annotator_id, domain)

        # Write with file locking
        try:
            wb = load_workbook(file_path)
            ws = wb.active

            malform_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row_data in rows:
                row = [
                    row_data.get('sample_id', ''),
                    row_data.get('text', '')[:500],  # Truncate text
                    row_data.get('raw_response', '')[:500],  # Truncate response
                    row_data.get('label', ''),
                    'YES' if row_data.get('malformed_flag', False) else 'NO',
                    row_data.get('parsing_error', ''),
                    row_data.get('validity_error', ''),
                    row_data.get('timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                ]

                ws.append(row)

                # Highlight malformed rows
                if row_data.get('malformed_flag', False):
                    row_num = ws.max_row
                    for col_num in range(1, len(self.HEADERS) + 1):
                        ws.cell(row_num, col_num).fill = malform_fill

            wb.save(file_path)
            logger.debug(f"Wrote {len(rows)} rows to {file_path.name}")

        except Exception as e:
            logger.error(f"Error writing to Excel file: {e}")
            raise

    def flush_buffer(self, annotator_id: int, domain: str) -> None:
        """
        Flush write buffer to Excel file.

        Args:
            annotator_id: Annotator ID
            domain: Domain name
        """
        worker_key = self._get_worker_key(annotator_id, domain)

        if worker_key not in self._buffers or not self._buffers[worker_key]:
            return

        rows = self._buffers[worker_key]
        self._buffers[worker_key] = []

        logger.info(f"Flushing buffer: {len(rows)} rows for {worker_key}")
        self.batch_write(annotator_id, domain, rows)

    def flush_all_buffers(self) -> None:
        """Flush all write buffers."""
        for worker_key in list(self._buffers.keys()):
            parts = worker_key.split('_')
            if len(parts) == 2:
                annotator_id = int(parts[0])
                domain = parts[1]
                self.flush_buffer(annotator_id, domain)

        logger.info("Flushed all buffers")

    def get_completed_sample_ids(self, annotator_id: int, domain: str) -> Set[str]:
        """
        Get set of completed sample IDs from Excel file.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Set of sample IDs
        """
        file_path = self._get_file_path(annotator_id, domain)

        if not file_path.exists():
            return set()

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            sample_ids = set()

            # Skip header row
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:  # Sample_ID column
                    sample_ids.add(str(row[0]))

            wb.close()
            logger.debug(f"Found {len(sample_ids)} completed samples in {file_path.name}")

            return sample_ids

        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            return set()

    def get_last_completed_sample_id(self, annotator_id: int, domain: str) -> Optional[str]:
        """
        Get last completed sample ID.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Last sample ID or None
        """
        file_path = self._get_file_path(annotator_id, domain)

        if not file_path.exists():
            return None

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            last_row = ws.max_row
            if last_row > 1:  # More than just header
                last_sample_id = ws.cell(last_row, 1).value
                wb.close()
                return str(last_sample_id) if last_sample_id else None

            wb.close()
            return None

        except Exception as e:
            logger.error(f"Error reading last sample ID: {e}")
            return None

    def get_progress(self, annotator_id: int, domain: str) -> Tuple[int, int]:
        """
        Get progress for worker.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Tuple of (completed, total)
        """
        # Get from Redis checkpoint
        completed, total = self.checkpoint_mgr.get_progress(annotator_id, domain)
        return (completed, total)

    def get_malformed_count(self, annotator_id: int, domain: str) -> int:
        """
        Get count of malformed annotations.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Malformed count
        """
        file_path = self._get_file_path(annotator_id, domain)

        if not file_path.exists():
            return 0

        try:
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            malformed_count = 0

            # Column E is Malformed_Flag (index 5)
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
                if row[0] == 'YES':
                    malformed_count += 1

            wb.close()
            return malformed_count

        except Exception as e:
            logger.error(f"Error counting malformed: {e}")
            return 0

    def sync_checkpoint_from_excel(self, annotator_id: int, domain: str) -> int:
        """
        Synchronize Redis checkpoint with Excel file state.

        This enables resume capability by reading existing Excel annotations
        and updating Redis checkpoint to match.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Number of samples synced
        """
        completed_ids = self.get_completed_sample_ids(annotator_id, domain)

        if not completed_ids:
            logger.debug(f"No completed samples found in Excel for {annotator_id}/{domain}")
            return 0

        # Update Redis checkpoint
        self.checkpoint_mgr.mark_completed_batch(annotator_id, domain, list(completed_ids))

        logger.info(f"Synced {len(completed_ids)} samples from Excel to checkpoint for {annotator_id}/{domain}")

        return len(completed_ids)

    def export_to_csv(self, annotator_id: int, domain: str, output_path: str) -> None:
        """
        Export Excel file to CSV.

        Args:
            annotator_id: Annotator ID
            domain: Domain name
            output_path: Output CSV file path
        """
        file_path = self._get_file_path(annotator_id, domain)

        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        try:
            import pandas as pd

            df = pd.read_excel(file_path)
            df.to_csv(output_path, index=False)

            logger.info(f"Exported {file_path.name} to CSV: {output_path}")

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    def get_file_info(self, annotator_id: int, domain: str) -> Dict:
        """
        Get information about Excel file.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Dictionary with file information
        """
        file_path = self._get_file_path(annotator_id, domain)

        if not file_path.exists():
            return {
                'exists': False,
                'path': str(file_path)
            }

        try:
            stats = file_path.stat()
            completed_ids = self.get_completed_sample_ids(annotator_id, domain)

            return {
                'exists': True,
                'path': str(file_path),
                'size_bytes': stats.st_size,
                'size_mb': round(stats.st_size / (1024 * 1024), 2),
                'created_at': datetime.fromtimestamp(stats.st_ctime).isoformat(),
                'modified_at': datetime.fromtimestamp(stats.st_mtime).isoformat(),
                'row_count': len(completed_ids),
                'malformed_count': self.get_malformed_count(annotator_id, domain)
            }

        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            return {
                'exists': True,
                'path': str(file_path),
                'error': str(e)
            }
