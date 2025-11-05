"""
Worker monitor for health checks, metrics collection, and auto-recovery.
"""
import os
import time
import psutil
import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from datetime import datetime, timedelta
import redis
from openpyxl import load_workbook

from ..core.checkpoint import RedisCheckpointManager
from ..storage.excel_manager import ExcelAnnotationManager


logger = logging.getLogger(__name__)


class WorkerMonitor:
    """
    Monitors worker health and collects metrics.

    Features:
    - Health checks (heartbeat, task completion, error rate, memory)
    - Excel file integrity verification
    - Stalled worker detection
    - Error worker detection
    - Auto-recovery mechanisms
    - System metrics collection
    """

    def __init__(self, redis_client: redis.Redis):
        """
        Initialize worker monitor.

        Args:
            redis_client: Redis client instance
        """
        self.redis = redis_client
        self.checkpoint_mgr = RedisCheckpointManager(redis_client)
        self.excel_mgr = ExcelAnnotationManager(
            output_dir='data/annotations',
            redis_client=redis_client
        )

        # Tracking for restart throttling
        self._restart_counts: Dict[str, List[datetime]] = {}
        self._max_restarts_per_hour = 3

        logger.info("WorkerMonitor initialized")

    # ═══════════════════════════════════════════════════════════
    # HELPER METHODS
    # ═══════════════════════════════════════════════════════════

    def _get_worker_key(self, annotator_id: int, domain: str) -> str:
        """Get unique key for worker."""
        return f"{annotator_id}_{domain}"

    def _get_redis_key(self, annotator_id: int, domain: str) -> str:
        """Get Redis key for worker metadata."""
        return f"worker:{annotator_id}:{domain}"

    def _get_metrics_key(self, annotator_id: int, domain: str) -> str:
        """Get Redis key for worker metrics."""
        return f"metrics:{annotator_id}:{domain}"

    # ═══════════════════════════════════════════════════════════
    # HEALTH CHECKS
    # ═══════════════════════════════════════════════════════════

    def check_worker_health(self, annotator_id: int, domain: str) -> Dict:
        """
        Perform comprehensive health check on a worker.

        Health criteria:
        - Heartbeat: Recent update within 60 seconds
        - Task completion: >0 tasks per minute expected
        - Error rate: <10% of total tasks
        - Memory: Worker process <500MB
        - Excel file: Accessible and not corrupted

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            Dictionary with health status
        """
        worker_key = self._get_worker_key(annotator_id, domain)
        redis_key = self._get_redis_key(annotator_id, domain)

        health = {
            'worker_key': worker_key,
            'healthy': True,
            'checks': {},
            'issues': [],
            'timestamp': datetime.now().isoformat()
        }

        # Get worker data
        worker_data = self.redis.hgetall(redis_key)

        if not worker_data:
            health['healthy'] = False
            health['issues'].append('Worker not registered')
            return health

        # Check 1: Heartbeat
        heartbeat_healthy = self._check_heartbeat(worker_data, health)

        # Check 2: Task completion rate
        completion_healthy = self._check_task_completion(annotator_id, domain, worker_data, health)

        # Check 3: Error rate
        error_rate_healthy = self._check_error_rate(annotator_id, domain, health)

        # Check 4: Memory usage
        memory_healthy = self._check_memory_usage(worker_data, health)

        # Check 5: Excel file integrity
        excel_healthy = self._check_excel_integrity(annotator_id, domain, health)

        # Overall health
        health['healthy'] = all([
            heartbeat_healthy,
            completion_healthy,
            error_rate_healthy,
            memory_healthy,
            excel_healthy
        ])

        return health

    def _check_heartbeat(self, worker_data: Dict, health: Dict) -> bool:
        """Check if heartbeat is recent (within 60 seconds)."""
        last_heartbeat = worker_data.get('last_heartbeat')

        if not last_heartbeat:
            health['checks']['heartbeat'] = 'FAIL'
            health['issues'].append('No heartbeat recorded')
            return False

        try:
            heartbeat_time = datetime.fromisoformat(last_heartbeat)
            age_seconds = (datetime.now() - heartbeat_time).total_seconds()

            if age_seconds > 60:
                health['checks']['heartbeat'] = 'FAIL'
                health['issues'].append(f'Heartbeat stale ({age_seconds:.0f}s old)')
                return False

            health['checks']['heartbeat'] = 'PASS'
            health['heartbeat_age_seconds'] = age_seconds
            return True

        except Exception as e:
            health['checks']['heartbeat'] = 'ERROR'
            health['issues'].append(f'Heartbeat parse error: {e}')
            return False

    def _check_task_completion(
        self,
        annotator_id: int,
        domain: str,
        worker_data: Dict,
        health: Dict
    ) -> bool:
        """Check task completion rate."""
        started_at = worker_data.get('started_at')
        processed_count = int(worker_data.get('processed_count', 0))

        if not started_at:
            health['checks']['completion_rate'] = 'UNKNOWN'
            return True  # Don't fail if we can't calculate

        try:
            start_time = datetime.fromisoformat(started_at)
            runtime_minutes = (datetime.now() - start_time).total_seconds() / 60.0

            if runtime_minutes < 1:
                # Too early to judge
                health['checks']['completion_rate'] = 'PENDING'
                return True

            tasks_per_minute = processed_count / runtime_minutes

            if tasks_per_minute == 0 and runtime_minutes > 5:
                # No tasks completed after 5 minutes
                health['checks']['completion_rate'] = 'FAIL'
                health['issues'].append('No tasks completed in 5+ minutes')
                return False

            health['checks']['completion_rate'] = 'PASS'
            health['tasks_per_minute'] = round(tasks_per_minute, 2)
            return True

        except Exception as e:
            health['checks']['completion_rate'] = 'ERROR'
            logger.warning(f"Error checking completion rate: {e}")
            return True  # Don't fail on calculation error

    def _check_error_rate(self, annotator_id: int, domain: str, health: Dict) -> bool:
        """Check error rate (<10% of total tasks)."""
        metrics_key = f"task_metrics:{annotator_id}:{domain}"
        metrics = self.redis.hgetall(metrics_key)

        if not metrics:
            health['checks']['error_rate'] = 'NO_DATA'
            return True

        total_tasks = int(metrics.get('total_tasks', 0))
        error_tasks = int(metrics.get('error_tasks', 0))
        malformed_tasks = int(metrics.get('malformed_tasks', 0))

        if total_tasks == 0:
            health['checks']['error_rate'] = 'NO_TASKS'
            return True

        error_count = error_tasks + malformed_tasks
        error_rate = (error_count / total_tasks) * 100

        if error_rate > 10:
            health['checks']['error_rate'] = 'FAIL'
            health['issues'].append(f'High error rate: {error_rate:.1f}%')
            health['error_rate_percent'] = error_rate
            return False

        health['checks']['error_rate'] = 'PASS'
        health['error_rate_percent'] = error_rate
        return True

    def _check_memory_usage(self, worker_data: Dict, health: Dict) -> bool:
        """Check if worker process memory usage is <500MB."""
        pid = worker_data.get('pid')

        if not pid:
            health['checks']['memory'] = 'NO_PID'
            return True

        try:
            process = psutil.Process(int(pid))
            memory_mb = process.memory_info().rss / (1024 * 1024)

            if memory_mb > 500:
                health['checks']['memory'] = 'FAIL'
                health['issues'].append(f'High memory usage: {memory_mb:.0f}MB')
                health['memory_mb'] = memory_mb
                return False

            health['checks']['memory'] = 'PASS'
            health['memory_mb'] = memory_mb
            return True

        except (psutil.NoSuchProcess, ValueError):
            health['checks']['memory'] = 'PROCESS_NOT_FOUND'
            health['issues'].append('Process not found')
            return False

        except Exception as e:
            health['checks']['memory'] = 'ERROR'
            logger.warning(f"Error checking memory: {e}")
            return True  # Don't fail on error

    def _check_excel_integrity(self, annotator_id: int, domain: str, health: Dict) -> bool:
        """Check if Excel file is accessible and not corrupted."""
        try:
            file_path = self.excel_mgr._get_file_path(annotator_id, domain)

            if not file_path.exists():
                health['checks']['excel'] = 'NO_FILE'
                health['issues'].append('Excel file does not exist')
                return False

            # Try to read first and last row to verify integrity
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active

            # Check if file has headers
            if ws.max_row < 1:
                health['checks']['excel'] = 'EMPTY'
                health['issues'].append('Excel file is empty')
                wb.close()
                return False

            # Try to read first data row
            if ws.max_row > 1:
                first_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))

            # Try to read last row
            if ws.max_row > 1:
                last_row = ws.cell(ws.max_row, 1).value

            wb.close()

            health['checks']['excel'] = 'PASS'
            health['excel_row_count'] = ws.max_row - 1  # Exclude header
            return True

        except Exception as e:
            health['checks']['excel'] = 'CORRUPTED'
            health['issues'].append(f'Excel file error: {str(e)[:100]}')
            return False

    # ═══════════════════════════════════════════════════════════
    # STATUS QUERIES
    # ═══════════════════════════════════════════════════════════

    def get_all_worker_statuses(self) -> Dict[str, Dict]:
        """
        Get status of all workers with health checks.

        Returns:
            Dictionary mapping worker keys to status info
        """
        statuses = {}
        pattern = "worker:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            annotator_id = int(parts[1])
            domain = parts[2]
            worker_key = self._get_worker_key(annotator_id, domain)

            # Get basic status
            worker_data = self.redis.hgetall(key)

            # Get progress
            completed, total = self.checkpoint_mgr.get_progress(annotator_id, domain)

            # Get health
            health = self.check_worker_health(annotator_id, domain)

            statuses[worker_key] = {
                'annotator_id': annotator_id,
                'domain': domain,
                'status': worker_data.get('status'),
                'pid': worker_data.get('pid'),
                'started_at': worker_data.get('started_at'),
                'last_heartbeat': worker_data.get('last_heartbeat'),
                'processed_count': int(worker_data.get('processed_count', 0)),
                'completed': completed,
                'total': total,
                'healthy': health['healthy'],
                'health_issues': health['issues']
            }

        return statuses

    def get_system_metrics(self) -> Dict:
        """
        Get system-level metrics (CPU, memory, Redis, disk).

        Returns:
            Dictionary with system metrics
        """
        metrics = {
            'timestamp': datetime.now().isoformat(),
            'cpu_percent': psutil.cpu_percent(interval=1.0),
            'memory': {
                'total_mb': psutil.virtual_memory().total / (1024 * 1024),
                'used_mb': psutil.virtual_memory().used / (1024 * 1024),
                'percent': psutil.virtual_memory().percent
            },
            'disk': {},
            'redis': {}
        }

        # Disk I/O for data directory
        try:
            data_path = Path('data')
            if data_path.exists():
                disk_usage = psutil.disk_usage(str(data_path))
                metrics['disk'] = {
                    'total_gb': disk_usage.total / (1024 ** 3),
                    'used_gb': disk_usage.used / (1024 ** 3),
                    'free_gb': disk_usage.free / (1024 ** 3),
                    'percent': disk_usage.percent
                }
        except Exception as e:
            logger.warning(f"Error getting disk metrics: {e}")

        # Redis stats
        try:
            redis_info = self.redis.info()
            metrics['redis'] = {
                'used_memory_mb': redis_info.get('used_memory', 0) / (1024 * 1024),
                'connected_clients': redis_info.get('connected_clients', 0),
                'total_commands_processed': redis_info.get('total_commands_processed', 0),
                'uptime_seconds': redis_info.get('uptime_in_seconds', 0)
            }
        except Exception as e:
            logger.warning(f"Error getting Redis metrics: {e}")

        return metrics

    # ═══════════════════════════════════════════════════════════
    # PROBLEM DETECTION
    # ═══════════════════════════════════════════════════════════

    def detect_stalled_workers(self, threshold_seconds: int = 60) -> List[Tuple[int, str]]:
        """
        Detect workers with no heartbeat for >threshold seconds.

        Args:
            threshold_seconds: Heartbeat age threshold

        Returns:
            List of (annotator_id, domain) tuples for stalled workers
        """
        stalled = []
        pattern = "worker:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            worker_data = self.redis.hgetall(key)
            if worker_data.get('status') != 'running':
                continue

            last_heartbeat = worker_data.get('last_heartbeat')
            if not last_heartbeat:
                annotator_id = int(parts[1])
                domain = parts[2]
                stalled.append((annotator_id, domain))
                continue

            try:
                heartbeat_time = datetime.fromisoformat(last_heartbeat)
                age_seconds = (datetime.now() - heartbeat_time).total_seconds()

                if age_seconds > threshold_seconds:
                    annotator_id = int(parts[1])
                    domain = parts[2]
                    stalled.append((annotator_id, domain))

            except Exception:
                continue

        if stalled:
            logger.warning(f"Detected {len(stalled)} stalled workers")

        return stalled

    def detect_error_workers(self, error_threshold: float = 20.0) -> List[Tuple[int, str]]:
        """
        Detect workers with high error rate (>threshold%).

        Args:
            error_threshold: Error rate threshold percentage

        Returns:
            List of (annotator_id, domain) tuples for error workers
        """
        error_workers = []
        pattern = "task_metrics:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            metrics = self.redis.hgetall(key)
            total_tasks = int(metrics.get('total_tasks', 0))

            if total_tasks < 10:
                # Need at least 10 tasks to judge
                continue

            error_tasks = int(metrics.get('error_tasks', 0))
            malformed_tasks = int(metrics.get('malformed_tasks', 0))

            error_count = error_tasks + malformed_tasks
            error_rate = (error_count / total_tasks) * 100

            if error_rate > error_threshold:
                annotator_id = int(parts[1])
                domain = parts[2]
                error_workers.append((annotator_id, domain))

        if error_workers:
            logger.warning(f"Detected {len(error_workers)} workers with high error rate")

        return error_workers

    # ═══════════════════════════════════════════════════════════
    # AUTO-RECOVERY
    # ═══════════════════════════════════════════════════════════

    def restart_stalled_workers(self, threshold_seconds: int = 60) -> int:
        """
        Detect and restart stalled workers with throttling.

        Args:
            threshold_seconds: Heartbeat age threshold

        Returns:
            Number of workers restarted
        """
        stalled = self.detect_stalled_workers(threshold_seconds)

        if not stalled:
            return 0

        restarted = 0

        for annotator_id, domain in stalled:
            # Check restart throttling
            if not self._can_restart_worker(annotator_id, domain):
                logger.warning(f"Restart throttled for {annotator_id}_{domain}")
                continue

            logger.warning(f"Auto-restarting stalled worker {annotator_id}_{domain}")

            # Mark worker for restart (actual restart done by launcher)
            redis_key = self._get_redis_key(annotator_id, domain)
            self.redis.hset(redis_key, 'status', 'restart_needed')
            self.redis.hset(redis_key, 'restart_reason', 'stalled')
            self.redis.hset(redis_key, 'restart_requested_at', datetime.now().isoformat())

            # Track restart
            self._record_restart(annotator_id, domain)

            restarted += 1

        if restarted > 0:
            logger.info(f"Marked {restarted} stalled workers for restart")

        return restarted

    def _can_restart_worker(self, annotator_id: int, domain: str) -> bool:
        """
        Check if worker can be restarted (not throttled).

        Max 3 restarts per hour per worker.

        Args:
            annotator_id: Annotator ID
            domain: Domain name

        Returns:
            True if restart allowed, False if throttled
        """
        worker_key = self._get_worker_key(annotator_id, domain)

        if worker_key not in self._restart_counts:
            return True

        # Filter restarts within last hour
        one_hour_ago = datetime.now() - timedelta(hours=1)
        recent_restarts = [
            t for t in self._restart_counts[worker_key]
            if t > one_hour_ago
        ]

        self._restart_counts[worker_key] = recent_restarts

        return len(recent_restarts) < self._max_restarts_per_hour

    def _record_restart(self, annotator_id: int, domain: str) -> None:
        """Record a worker restart for throttling."""
        worker_key = self._get_worker_key(annotator_id, domain)

        if worker_key not in self._restart_counts:
            self._restart_counts[worker_key] = []

        self._restart_counts[worker_key].append(datetime.now())

    # ═══════════════════════════════════════════════════════════
    # EXCEL FILE MONITORING
    # ═══════════════════════════════════════════════════════════

    def verify_excel_integrity(self) -> Dict[str, bool]:
        """
        Verify integrity of all Excel files.

        Returns:
            Dictionary mapping worker keys to integrity status
        """
        results = {}
        pattern = "worker:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            annotator_id = int(parts[1])
            domain = parts[2]
            worker_key = self._get_worker_key(annotator_id, domain)

            try:
                file_path = self.excel_mgr._get_file_path(annotator_id, domain)

                if not file_path.exists():
                    results[worker_key] = False
                    continue

                # Try to open and read
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active

                # Check basic structure
                if ws.max_row >= 1:
                    # File is readable
                    results[worker_key] = True
                else:
                    results[worker_key] = False

                wb.close()

            except Exception as e:
                logger.error(f"Excel integrity check failed for {worker_key}: {e}")
                results[worker_key] = False

        return results

    def get_excel_file_sizes(self) -> Dict[str, int]:
        """
        Get file sizes for all Excel files.

        Returns:
            Dictionary mapping worker keys to file sizes in bytes
        """
        sizes = {}
        pattern = "worker:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            annotator_id = int(parts[1])
            domain = parts[2]
            worker_key = self._get_worker_key(annotator_id, domain)

            try:
                file_path = self.excel_mgr._get_file_path(annotator_id, domain)

                if file_path.exists():
                    sizes[worker_key] = file_path.stat().st_size

                    # Alert on large files (>100MB)
                    size_mb = sizes[worker_key] / (1024 * 1024)
                    if size_mb > 100:
                        logger.warning(f"Large Excel file for {worker_key}: {size_mb:.1f}MB")

            except Exception as e:
                logger.error(f"Error getting file size for {worker_key}: {e}")

        return sizes

    # ═══════════════════════════════════════════════════════════
    # METRICS COLLECTION
    # ═══════════════════════════════════════════════════════════

    def collect_worker_metrics(self, annotator_id: int, domain: str) -> None:
        """
        Collect and aggregate metrics for a worker.

        Stores in Redis: metrics:{annotator_id}:{domain}

        Args:
            annotator_id: Annotator ID
            domain: Domain name
        """
        metrics_key = self._get_metrics_key(annotator_id, domain)
        task_metrics_key = f"task_metrics:{annotator_id}:{domain}"

        # Get task metrics
        task_metrics = self.redis.hgetall(task_metrics_key)

        if not task_metrics:
            return

        total_tasks = int(task_metrics.get('total_tasks', 0))
        successful_tasks = int(task_metrics.get('successful_tasks', 0))
        malformed_tasks = int(task_metrics.get('malformed_tasks', 0))
        error_tasks = int(task_metrics.get('error_tasks', 0))
        total_duration = float(task_metrics.get('total_duration', 0.0))

        # Calculate metrics
        avg_task_duration = total_duration / successful_tasks if successful_tasks > 0 else 0
        error_rate = ((malformed_tasks + error_tasks) / total_tasks * 100) if total_tasks > 0 else 0

        # Get Excel file size
        try:
            file_path = self.excel_mgr._get_file_path(annotator_id, domain)
            excel_size = file_path.stat().st_size if file_path.exists() else 0
        except Exception:
            excel_size = 0

        # Store aggregated metrics
        aggregated_metrics = {
            'tasks_completed': successful_tasks,
            'tasks_failed': error_tasks,
            'tasks_malformed': malformed_tasks,
            'avg_task_duration': avg_task_duration,
            'error_rate': error_rate,
            'excel_file_size': excel_size,
            'last_updated': datetime.now().isoformat()
        }

        self.redis.hset(metrics_key, mapping=aggregated_metrics)
        self.redis.expire(metrics_key, 86400)  # 24 hours

    def collect_all_metrics(self) -> None:
        """Collect metrics for all workers."""
        pattern = "worker:*"
        keys = self.redis.keys(pattern)

        for key in keys:
            parts = key.split(':')
            if len(parts) != 3:
                continue

            annotator_id = int(parts[1])
            domain = parts[2]

            self.collect_worker_metrics(annotator_id, domain)

        logger.debug(f"Collected metrics for {len(keys)} workers")
