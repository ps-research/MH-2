"""
Tests for Excel annotation manager.
"""
import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import redis
from openpyxl import Workbook

from src.storage.excel_manager import ExcelAnnotationManager


class TestExcelAnnotationManager:
    """Tests for Excel annotation manager."""

    @pytest.fixture
    def redis_mock(self):
        """Create mock Redis client."""
        mock_redis = Mock(spec=redis.Redis)
        mock_redis.hgetall.return_value = {}
        mock_redis.sadd.return_value = True
        mock_redis.sismember.return_value = False
        return mock_redis

    @pytest.fixture
    def temp_output_dir(self, tmp_path):
        """Create temporary output directory."""
        output_dir = tmp_path / "annotations"
        output_dir.mkdir()
        return output_dir

    @pytest.fixture
    def excel_manager(self, temp_output_dir, redis_mock):
        """Create Excel manager instance."""
        return ExcelAnnotationManager(
            output_dir=str(temp_output_dir),
            redis_client=redis_mock,
            buffer_size=10
        )

    def test_initialization(self, excel_manager, temp_output_dir):
        """Test Excel manager initialization."""
        assert excel_manager.output_dir == Path(temp_output_dir)
        assert excel_manager.buffer_size == 10

    def test_get_file_path(self, excel_manager):
        """Test file path generation."""
        file_path = excel_manager._get_file_path(1, 'urgency')

        assert file_path.name == 'annotator_1_urgency.xlsx'
        assert file_path.parent == excel_manager.output_dir

    def test_initialize_file(self, excel_manager):
        """Test Excel file initialization."""
        file_path_str = excel_manager.initialize_file(1, 'urgency')
        file_path = Path(file_path_str)

        assert file_path.exists()
        assert file_path.suffix == '.xlsx'

        # Verify file structure
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert headers == excel_manager.HEADERS

        wb.close()

    def test_write_annotation_buffered(self, excel_manager):
        """Test buffered annotation writing."""
        # Initialize file
        excel_manager.initialize_file(1, 'urgency')

        # Write annotation (should be buffered)
        row_data = {
            'sample_id': 'TEST-001',
            'text': 'Test text',
            'raw_response': 'Test response <<LEVEL_3>>',
            'label': 'LEVEL_3',
            'malformed_flag': False,
            'parsing_error': '',
            'validity_error': '',
            'timestamp': '2025-01-01 12:00:00'
        }

        excel_manager.write_annotation(1, 'urgency', row_data)

        # Check buffer
        worker_key = excel_manager._get_worker_key(1, 'urgency')
        assert len(excel_manager._buffers[worker_key]) == 1

    def test_batch_write(self, excel_manager):
        """Test batch writing to Excel."""
        # Initialize file
        excel_manager.initialize_file(1, 'urgency')

        # Prepare batch data
        rows = [
            {
                'sample_id': f'TEST-{i:03d}',
                'text': f'Test text {i}',
                'raw_response': f'Response {i} <<LEVEL_2>>',
                'label': 'LEVEL_2',
                'malformed_flag': False,
                'parsing_error': '',
                'validity_error': '',
                'timestamp': '2025-01-01 12:00:00'
            }
            for i in range(5)
        ]

        excel_manager.batch_write(1, 'urgency', rows)

        # Verify file has rows
        file_path = excel_manager._get_file_path(1, 'urgency')
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb.active

        # Should have header + 5 data rows
        assert ws.max_row == 6
        wb.close()

    def test_get_completed_sample_ids(self, excel_manager):
        """Test retrieving completed sample IDs."""
        # Initialize and write data
        excel_manager.initialize_file(1, 'urgency')

        rows = [
            {
                'sample_id': f'TEST-{i:03d}',
                'text': f'Test text {i}',
                'raw_response': 'Response',
                'label': 'LEVEL_2',
                'malformed_flag': False,
                'parsing_error': '',
                'validity_error': '',
                'timestamp': '2025-01-01 12:00:00'
            }
            for i in range(3)
        ]

        excel_manager.batch_write(1, 'urgency', rows)

        # Get completed IDs
        completed_ids = excel_manager.get_completed_sample_ids(1, 'urgency')

        assert len(completed_ids) == 3
        assert 'TEST-000' in completed_ids
        assert 'TEST-001' in completed_ids
        assert 'TEST-002' in completed_ids

    def test_get_malformed_count(self, excel_manager):
        """Test counting malformed responses."""
        excel_manager.initialize_file(1, 'urgency')

        rows = [
            {
                'sample_id': 'TEST-001',
                'text': 'Text 1',
                'raw_response': 'Response 1',
                'label': 'LEVEL_2',
                'malformed_flag': False,
                'parsing_error': '',
                'validity_error': '',
                'timestamp': '2025-01-01 12:00:00'
            },
            {
                'sample_id': 'TEST-002',
                'text': 'Text 2',
                'raw_response': 'Response 2',
                'label': '',
                'malformed_flag': True,
                'parsing_error': 'Parse error',
                'validity_error': '',
                'timestamp': '2025-01-01 12:00:00'
            }
        ]

        excel_manager.batch_write(1, 'urgency', rows)

        malformed_count = excel_manager.get_malformed_count(1, 'urgency')
        assert malformed_count == 1

    def test_flush_buffer(self, excel_manager):
        """Test buffer flushing."""
        excel_manager.initialize_file(1, 'urgency')

        # Add to buffer
        for i in range(5):
            row_data = {
                'sample_id': f'TEST-{i:03d}',
                'text': f'Test text {i}',
                'raw_response': 'Response',
                'label': 'LEVEL_2',
                'malformed_flag': False,
                'parsing_error': '',
                'validity_error': '',
                'timestamp': '2025-01-01 12:00:00'
            }
            excel_manager.write_annotation(1, 'urgency', row_data)

        # Flush
        excel_manager.flush_buffer(1, 'urgency')

        # Buffer should be empty
        worker_key = excel_manager._get_worker_key(1, 'urgency')
        assert len(excel_manager._buffers[worker_key]) == 0

        # File should have data
        completed_ids = excel_manager.get_completed_sample_ids(1, 'urgency')
        assert len(completed_ids) == 5


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
